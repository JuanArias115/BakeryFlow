#!/usr/bin/env python3
from __future__ import annotations

import argparse
import math
import re
import sys
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
import xlrd
from openpyxl import load_workbook


UNIT_SEED = [
    {"name": "gramo", "abbreviation": "g", "type": "masa"},
    {"name": "kilogramo", "abbreviation": "kg", "type": "masa"},
    {"name": "mililitro", "abbreviation": "ml", "type": "volumen"},
    {"name": "litro", "abbreviation": "l", "type": "volumen"},
    {"name": "unidad", "abbreviation": "und", "type": "unidad"},
    {"name": "caja", "abbreviation": "caja", "type": "empaque"},
    {"name": "metro", "abbreviation": "m", "type": "longitud"},
]

PRODUCT_KEYWORDS = {
    "cake": ["torta", "cake"],
    "cheesecake": ["cheescake", "cheesecake", "chescake"],
    "brownie": ["brownie", "browniw"],
    "cupcake": ["cupcake", "muffin"],
    "alfajor": ["alfajor"],
    "trufa": ["trufa"],
    "desayuno": ["desayuno", "pancake", "huevos"],
    "galleta": ["galleta", "polvoron"],
    "tres_leches": ["tres leches"],
    "vasca": ["vasca"],
    "flores": ["flores", "floristeria", "arreglo"],
    "gelatina": ["gelatina"],
    "sandwich": ["sandwich", "tortilla", "pan"],
    "corazon": ["corazon", "corazón"],
}

CUSTOMER_NOISE = {
    "brownie",
    "cupcake",
    "cupcakes",
    "galletas",
    "flores",
    "chescake",
    "cheescake",
    "desayuno",
    "alfajores x12",
    "alfajores",
    "torta pequeñna",
    "torta pequena",
    "totrta pequena",
    "toratas y galleta",
    "torta niños",
    "torta fresa",
}


@dataclass
class ProductRecipe:
    category: str
    name: str
    sale_price: float
    details: list[dict[str, Any]]
    notes: str


class ApiClient:
    def __init__(self, base_url: str, email: str, password: str):
        self.base_url = base_url.rstrip("/")
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        response = self.session.post(
            f"{self.base_url}/auth/login",
            json={"email": email, "password": password},
            timeout=30,
        )
        response.raise_for_status()
        token = response.json()["data"]["token"]
        self.session.headers.update({"Authorization": f"Bearer {token}"})

    def get(self, endpoint: str, **params: Any) -> Any:
        response = self.session.get(f"{self.base_url}/{endpoint.lstrip('/')}", params=params, timeout=60)
        response.raise_for_status()
        return response.json()["data"]

    def post(self, endpoint: str, payload: dict[str, Any]) -> Any:
        response = self.session.post(f"{self.base_url}/{endpoint.lstrip('/')}", json=payload, timeout=60)
        response.raise_for_status()
        return response.json()["data"]

    def list_all(self, endpoint: str) -> list[dict[str, Any]]:
        page = 1
        items: list[dict[str, Any]] = []
        while True:
            data = self.get(endpoint, page=page, pageSize=100)
            page_items = data.get("items", [])
            items.extend(page_items)
            if len(items) >= data.get("totalCount", 0) or not page_items:
                return items
            page += 1


def slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    normalized = normalized.lower().strip()
    normalized = re.sub(r"[^a-z0-9]+", "-", normalized)
    return normalized.strip("-")


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def normalize_name(value: str) -> str:
    text = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return normalize_spaces(text).lower()


def clean_ingredient_name(raw: str) -> str:
    text = normalize_spaces(str(raw).upper())
    text = re.sub(r"UTILIDADES.*", "UTILIDADES", text)
    text = re.sub(r"\bX\s*\d+[.,]?\d*\s*(GR|G|ML|UND|UN|KG|KILO|LIBRA|LB|MT)\b", "", text)
    text = re.sub(r"\bBAR+RA\s*X?\s*\d+[A-Z]*\b", "", text)
    text = re.sub(r"\bBANDEJA\b", "", text)
    text = re.sub(r"\bPQT\b.*", "", text)
    text = re.sub(r"\bBOTELLA\b.*", "", text)
    text = re.sub(r"\b300GR\b", "", text)
    text = re.sub(r"\b250 ML\b", "", text)
    text = text.replace("  ", " ").strip(" -")

    aliases = {
        "ESCENCIA VAINILLA": "ESENCIA DE VAINILLA",
        "ESCENCIA VAINILLA 60 ML": "ESENCIA DE VAINILLA",
        "HUEVO": "HUEVO",
        "HUEVOS": "HUEVO",
        "HUEVO X 30": "HUEVO",
        "HUEVO BANDEJA": "HUEVO",
        "LECHE X 900 ML": "LECHE",
        "LECHE ENTERA": "LECHE ENTERA",
        "MANTEQUILLA X 125 GR": "MANTEQUILLA",
        "AZUCAR X KILO": "AZUCAR",
        "AZUCAR MORENA X KILO": "AZUCAR MORENA",
        "HARINA X LIBRA": "HARINA",
        "POLVO HORNEAR X LIBRA": "POLVO DE HORNEAR",
        "POLVO HORNEAR": "POLVO DE HORNEAR",
        "LECHE CONDENSADA ": "LECHE CONDENSADA",
        "CREMA DE LECHE ": "CREMA DE LECHE",
        "QUESO CREMA 250 ML": "QUESO CREMA",
        "EMPAQUE TRIANGULO": "EMPAQUE TRIANGULO",
        "DECORADO COLORANTE CHIPS": "DECORADO",
        "AREQUIPE + OREO/ CHOCOLATE+ CHIPS": "TOPPING MIXTO",
        "DECORADA": "DECORADA",
        "CAJA 3.000(": "CAJA",
        "CAJA": "CAJA",
        "BASE": "BASE",
        "FRUTA/ TOPPING": "FRUTA O TOPPING",
        "LECHE ALMIBAR XA MOJAR LECHE 3,2+COND 15+CREM 10": "LECHE ALMIBAR",
        "RELLENO INTERMEDIO NUTELL 28,6 AREQUI 16": "RELLENO INTERMEDIO",
        "CHOCOLATE+ CHIPS": "CHOCOLATE Y CHIPS",
    }
    return normalize_spaces(aliases.get(text, text.title()))


def infer_unit(raw_name: str, presentation: str | None = None) -> str:
    text = f"{raw_name} {presentation or ''}".upper()
    if "MT" in text or "METRO" in text:
        return "metro"
    if any(token in text for token in ["ML", "LITRO", "LT"]):
        return "mililitro"
    if any(token in text for token in ["KILO", "KG", "LIBRA", "LB", "GR", "G "]):
        return "gramo"
    if any(token in text for token in ["UND", "UNIDAD", "BANDEJA", "PQT", "BOTELLA", "CAJA", "X 30"]):
        return "unidad"
    return "unidad"


def num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value):
            return None
        return float(value)
    if isinstance(value, str):
        value = value.strip().replace(",", ".")
        if not value:
            return None
        try:
            return float(value)
        except ValueError:
            return None
    return None


def text(value: Any) -> str:
    if value is None:
        return ""
    return normalize_spaces(str(value))


def iso_utc(dt: datetime) -> str:
    return dt.replace(tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")


MONTHS = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}


def sheet_fallback_date(sheet_name: str, row_index: int) -> str:
    month = MONTHS.get(sheet_name.strip().upper(), datetime.now(timezone.utc).month)
    day = max(1, min(28, row_index - 4))
    return iso_utc(datetime(2026, month, day, 12, 0, 0))


def xls_date(value: Any, datemode: int) -> str | None:
    numeric = num(value)
    if numeric is None:
        return None
    dt = xlrd.xldate_as_datetime(numeric, datemode)
    return iso_utc(datetime(dt.year, dt.month, dt.day, 12, 0, 0))


class Importer:
    def __init__(self, client: ApiClient):
        self.client = client
        self.units_by_name: dict[str, dict[str, Any]] = {}
        self.categories_by_name: dict[str, dict[str, Any]] = {}
        self.ingredients_by_name: dict[str, dict[str, Any]] = {}
        self.products_by_name: dict[str, dict[str, Any]] = {}
        self.suppliers_by_name: dict[str, dict[str, Any]] = {}
        self.customers_by_name: dict[str, dict[str, Any]] = {}

    def refresh(self) -> None:
        self.units_by_name = {normalize_name(item["name"]): item for item in self.client.list_all("units")}
        self.categories_by_name = {normalize_name(item["name"]): item for item in self.client.list_all("categories")}
        self.ingredients_by_name = {normalize_name(item["name"]): item for item in self.client.list_all("ingredients")}
        self.products_by_name = {normalize_name(item["name"]): item for item in self.client.list_all("products")}
        self.suppliers_by_name = {normalize_name(item["name"]): item for item in self.client.list_all("suppliers")}
        self.customers_by_name = {normalize_name(item["name"]): item for item in self.client.list_all("customers")}

    def ensure_unit(self, name: str, abbr: str, unit_type: str) -> dict[str, Any]:
        key = normalize_name(name)
        if key in self.units_by_name:
            return self.units_by_name[key]
        created = self.client.post("units", {"name": name, "abbreviation": abbr, "type": unit_type, "isActive": True})
        self.units_by_name[key] = created
        return created

    def ensure_category(self, name: str, description: str | None = None) -> dict[str, Any]:
        key = normalize_name(name)
        if key in self.categories_by_name:
            return self.categories_by_name[key]
        created = self.client.post("categories", {"name": name, "description": description, "isActive": True})
        self.categories_by_name[key] = created
        return created

    def ensure_supplier(self, name: str, notes: str | None = None) -> dict[str, Any]:
        key = normalize_name(name)
        if key in self.suppliers_by_name:
            return self.suppliers_by_name[key]
        created = self.client.post(
            "suppliers",
            {"name": name, "phone": None, "email": None, "address": None, "contact": None, "notes": notes, "isActive": True},
        )
        self.suppliers_by_name[key] = created
        return created

    def ensure_customer(self, name: str, notes: str | None = None) -> dict[str, Any]:
        key = normalize_name(name)
        if key in self.customers_by_name:
            return self.customers_by_name[key]
        created = self.client.post(
            "customers",
            {"name": name, "phone": None, "email": None, "address": None, "notes": notes, "isActive": True},
        )
        self.customers_by_name[key] = created
        return created

    def ensure_ingredient(self, name: str, unit_name: str, average_cost: float, description: str | None = None) -> dict[str, Any]:
        key = normalize_name(name)
        existing = self.ingredients_by_name.get(key)
        if existing:
            return existing

        unit = self.units_by_name[normalize_name(unit_name)]
        payload = {
            "code": None,
            "name": name,
            "unitOfMeasureId": unit["id"],
            "stockCurrent": 0,
            "stockMinimum": 0,
            "averageCost": round(max(average_cost, 0), 4),
            "description": description,
            "isActive": True,
        }
        created = self.client.post("ingredients", payload)
        self.ingredients_by_name[key] = created
        return created

    def ensure_product(self, category_name: str, name: str, sale_price: float, description: str | None = None) -> dict[str, Any]:
        key = normalize_name(name)
        existing = self.products_by_name.get(key)
        if existing:
            return existing
        category = self.ensure_category(category_name)
        payload = {
            "code": None,
            "name": name,
            "categoryId": category["id"],
            "unitSale": "unidad",
            "salePrice": round(max(sale_price, 1), 2),
            "description": description,
            "isActive": True,
        }
        created = self.client.post("products", payload)
        self.products_by_name[key] = created
        return created

    def ensure_recipe(self, product: dict[str, Any], details: list[dict[str, Any]], notes: str) -> None:
        existing = self.client.get("recipes", page=1, pageSize=100, search=product["name"])
        for item in existing.get("items", []):
            if item["productId"] == product["id"]:
                return
        payload = {
            "productId": product["id"],
            "yield": 1,
            "yieldUnit": "unidad",
            "packagingCost": 0,
            "notes": notes,
            "isActive": True,
            "details": details,
        }
        self.client.post("recipes", payload)


STANDARD_BLOCKS = [
    {
        "sheet": "CAKES",
        "category": "Cakes",
        "header_row": 1,
        "unit_cost_col": 3,
        "product_map": {
            5: {"name": "Cake personal", "qty_col": 4, "price_row": 19},
            6: {"name": "Cake personal decorada", "qty_col": 4, "price_row": 19},
            8: {"name": "Cake 4 a 6", "qty_col": 7, "price_row": 19},
            9: {"name": "Cake 4 a 6 decorada", "qty_col": 7, "price_row": 19},
            11: {"name": "Cake 12 a 14", "qty_col": 10, "price_row": 19},
            12: {"name": "Cake 12 a 14 decorada", "qty_col": 10, "price_row": 19},
        },
    },
    {
        "sheet": "CHEESECAKES",
        "category": "Cheesecakes",
        "header_row": 1,
        "unit_cost_col": 4,
        "presentation_col": 3,
        "product_map": {
            6: {"name": "Cheesecake personal", "qty_col": 5, "price_row": 14},
            8: {"name": "Cheesecake corazón", "qty_col": 7, "price_row": 14},
            10: {"name": "Cheesecake mediano", "qty_col": 9, "price_row": 14},
            12: {"name": "Cheesecake grande", "qty_col": 11, "price_row": 14},
        },
    },
    {
        "sheet": "BROWNIES",
        "category": "Brownies",
        "header_row": 1,
        "unit_cost_col": 3,
        "product_map": {
            5: {"name": "Brownie clásico", "qty_col": 4, "price_row": 15},
            7: {"name": "Brownie mega corazón", "qty_col": 6, "price_row": 15},
            9: {"name": "Brownie cuadrado", "qty_col": 8, "price_row": 15},
        },
    },
    {
        "sheet": "ALFAJORES",
        "category": "Alfajores",
        "header_row": 1,
        "unit_cost_col": 3,
        "product_map": {
            5: {"name": "Alfajores x12", "qty_col": 4, "price_value": 24000},
        },
    },
    {
        "sheet": "zanahoria",
        "category": "Torta de zanahoria",
        "header_row": 1,
        "unit_cost_col": 3,
        "product_map": {
            5: {"name": "Torta de zanahoria media", "qty_col": 4, "price_row": 14},
            7: {"name": "Torta de zanahoria completa", "qty_col": 6, "price_row": 14},
        },
    },
    {
        "sheet": "CUPCAKES",
        "category": "Cupcakes",
        "header_row": 1,
        "unit_cost_col": 3,
        "product_map": {
            5: {"name": "Cupcakes x6", "qty_col": 4, "price_row": 16},
            7: {"name": "Cupcakes x12", "qty_col": 6, "price_row": 16},
        },
    },
    {
        "sheet": "DESAYUNOS",
        "category": "Desayunos",
        "header_row": 1,
        "unit_cost_col": 4,
        "product_map": {
            6: {"name": "Desayuno especial", "qty_col": 5, "price_value": 52000},
        },
    },
    {
        "sheet": "TRUFAS",
        "category": "Trufas",
        "header_row": 1,
        "unit_cost_col": 3,
        "product_map": {
            5: {"name": "Trufas x6", "qty_col": 4, "price_value": 17424},
            7: {"name": "Trufas x12", "qty_col": 6, "price_value": 26000},
        },
    },
    {
        "sheet": "GALLETAS",
        "category": "Galletas",
        "header_row": 1,
        "unit_cost_col": 3,
        "product_map": {
            5: {"name": "Galletas x10", "qty_col": 4, "price_row": 19},
        },
    },
    {
        "sheet": "tres leches 10und",
        "category": "Tres leches",
        "header_row": 1,
        "unit_cost_col": 3,
        "product_map": {
            5: {"name": "Tres leches x10", "qty_col": 4, "price_row": 22},
        },
    },
    {
        "sheet": "vasca",
        "category": "Vasca",
        "header_row": 1,
        "unit_cost_col": 3,
        "product_map": {
            5: {"name": "Tarta vasca", "qty_col": 4, "price_row": 15},
        },
    },
]

AGGREGATE_PRODUCTS = [
    {"sheet": "SERGIO ", "category": "Gelatinas", "name": "Gelatina x4", "unit_cost": 1912.15, "sale_price": 4700},
    {"sheet": "SERGIO ", "category": "Crookies", "name": "Crookies", "unit_cost": 644.1111, "sale_price": 2100},
    {"sheet": "SERGIO ", "category": "Alfajores", "name": "Alfajores x8", "unit_cost": 586.375, "sale_price": 2000},
    {"sheet": "SERGIO ", "category": "Sandwich", "name": "Sandwich pan", "unit_cost": 2550, "sale_price": 3900},
    {"sheet": "SERGIO ", "category": "Sandwich", "name": "Sandwich tortilla", "unit_cost": 5016, "sale_price": 7500},
    {"sheet": "SERGIO ", "category": "Cupcakes", "name": "Muffins x6", "unit_cost": 10762.3833, "sale_price": 18000},
    {"sheet": "SERGIO ", "category": "Cupcakes", "name": "Muffins x12", "unit_cost": 17924.7666, "sale_price": 36000},
    {"sheet": "SERGIO ", "category": "Corazones", "name": "Corazones", "unit_cost": 748, "sale_price": 1800},
    {"sheet": "flores", "category": "Flores", "name": "Arreglo flores", "unit_cost": 70000, "sale_price": 110000},
    {"sheet": "flores", "category": "Flores", "name": "Arreglo 2 flores", "unit_cost": 92000, "sale_price": 150000},
    {"sheet": "GALLETAS", "category": "Galletas", "name": "Caja navideña", "unit_cost": 30855.7731, "sale_price": 40000},
    {"sheet": "GALLETAS", "category": "Galletas", "name": "Caja farol", "unit_cost": 8840, "sale_price": 17680},
    {"sheet": "GALLETAS", "category": "Galletas", "name": "Bolsa mantequilla", "unit_cost": 3892.9552, "sale_price": 11678.87},
    {"sheet": "tres leches 10und", "category": "Tres leches", "name": "Tres leches unidad", "unit_cost": 2176.4056, "sale_price": 7500},
    {"sheet": "vasca", "category": "Vasca", "name": "Tarta vasca por unidad", "unit_cost": 2207.1787, "sale_price": 7500},
]


def build_standard_product_recipes(path: Path, importer: Importer) -> list[ProductRecipe]:
    workbook = load_workbook(path, data_only=True)
    recipes: list[ProductRecipe] = []

    for config in STANDARD_BLOCKS:
        sheet = workbook[config["sheet"]]
        ingredient_cache: dict[int, dict[str, Any]] = {}

        max_row = sheet.max_row
        for row_index in range(config["header_row"] + 1, max_row + 1):
            first_cell = text(sheet.cell(row_index, 1).value)
            if not first_cell:
                continue
            upper_first = first_cell.upper()
            if upper_first.startswith(("COSTO", "PRECIO", "VENTA", "TOTAL", "UNITARIO")):
                continue

            unit_cost = num(sheet.cell(row_index, config["unit_cost_col"]).value)
            if unit_cost is None:
                continue

            presentation = text(sheet.cell(row_index, config.get("presentation_col", 0)).value) if config.get("presentation_col") else None
            ingredient_name = clean_ingredient_name(first_cell)
            unit_name = infer_unit(first_cell, presentation)
            ingredient = importer.ensure_ingredient(
                ingredient_name,
                unit_name,
                unit_cost,
                f"Importado desde hoja {config['sheet']}: {first_cell}",
            )
            ingredient_cache[row_index] = ingredient

        for product_col, product_config in config["product_map"].items():
            price = product_config.get("price_value")
            if price is None and product_config.get("price_row"):
                price = num(sheet.cell(product_config["price_row"], product_col).value)
            if price is None or price <= 0:
                continue

            details: list[dict[str, Any]] = []
            for row_index, ingredient in ingredient_cache.items():
                quantity = num(sheet.cell(row_index, product_config["qty_col"]).value)
                if quantity is None or quantity <= 0:
                    continue
                details.append(
                    {
                        "ingredientId": ingredient["id"],
                        "quantity": round(quantity, 4),
                        "unitOfMeasureId": ingredient["unitOfMeasureId"],
                    }
                )

            if not details:
                continue

            recipes.append(
                ProductRecipe(
                    category=config["category"],
                    name=product_config["name"],
                    sale_price=float(price),
                    details=details,
                    notes=f"Importado desde hoja {config['sheet']}",
                )
            )

    return recipes


def ensure_aggregate_products(importer: Importer) -> None:
    cost_zero = importer.ensure_ingredient("Costo base importado", "unidad", 0, "Ingrediente técnico para recetas históricas.")
    for item in AGGREGATE_PRODUCTS:
        ingredient = importer.ensure_ingredient(
            f"Costo base {item['name']}",
            "unidad",
            float(item["unit_cost"]),
            f"Costo agregado importado desde {item['sheet']}.",
        )
        product = importer.ensure_product(
            item["category"],
            item["name"],
            float(item["sale_price"]),
            f"Producto importado desde hoja {item['sheet']}.",
        )
        importer.ensure_recipe(
            product,
            [{"ingredientId": ingredient["id"], "quantity": 1, "unitOfMeasureId": ingredient["unitOfMeasureId"]}],
            f"Receta agregada importada desde {item['sheet']}.",
        )

    generic_sale = importer.ensure_product("Histórico", "Venta histórica importada", 1000, "Producto técnico para ventas resumidas.")
    importer.ensure_recipe(
        generic_sale,
        [{"ingredientId": cost_zero["id"], "quantity": 1, "unitOfMeasureId": cost_zero["unitOfMeasureId"]}],
        "Receta técnica para ventas históricas importadas.",
    )
    importer.ensure_ingredient("Compra histórica importada", "unidad", 0, "Ingrediente técnico para compras resumidas.")


def seed_units(importer: Importer) -> None:
    for unit in UNIT_SEED:
        importer.ensure_unit(unit["name"], unit["abbreviation"], unit["type"])


def import_products_and_recipes(importer: Importer, cost_workbook: Path) -> None:
    recipes = build_standard_product_recipes(cost_workbook, importer)
    created = 0
    for recipe in recipes:
        product = importer.ensure_product(recipe.category, recipe.name, recipe.sale_price, recipe.notes)
        importer.ensure_recipe(product, recipe.details, recipe.notes)
        created += 1
    ensure_aggregate_products(importer)
    print(f"Productos/recetas procesados: {created + len(AGGREGATE_PRODUCTS) + 1}")


def extract_purchase_rows(path: Path) -> list[dict[str, Any]]:
    book = xlrd.open_workbook(str(path))
    rows: list[dict[str, Any]] = []
    for sheet_name in book.sheet_names():
        if sheet_name.upper() == "TOTALES":
            continue
        sheet = book.sheet_by_name(sheet_name)
        for row_idx in range(5, sheet.nrows):
            row = sheet.row_values(row_idx)
            supplier = text(row[2] if len(row) > 2 else "")
            order_text = text(row[3] if len(row) > 3 else "")
            invoice = text(row[0] if len(row) > 0 else "")
            total = next((num(row[index]) for index in [9, 8, 7, 6] if index < len(row) and num(row[index]) and num(row[index]) > 0), None)
            if not supplier or total is None or total <= 0:
                continue
            rows.append(
                {
                    "sheet": sheet_name,
                    "supplier": supplier.title(),
                    "invoice": invoice or None,
                    "date": xls_date(row[1] if len(row) > 1 else None, book.datemode) or sheet_fallback_date(sheet_name, row_idx),
                    "notes": normalize_spaces(f"Importado desde compras {sheet_name}. Pedido: {order_text}")[:500],
                    "description": (order_text or supplier)[:200],
                    "total": round(total, 2),
                }
            )
    return rows


def probable_customer(name: str) -> bool:
    key = normalize_name(name)
    if not key or key in CUSTOMER_NOISE:
        return False
    return not any(word in key for word in ["torta", "brownie", "cupcake", "galleta", "cheescake", "chescake", "flores"])


def match_product(products_by_name: dict[str, dict[str, Any]], blob: str) -> dict[str, Any] | None:
    text_blob = normalize_name(blob)
    if not text_blob:
        return None
    checks = [
        ("cupcakes x12", ["cupcake", "x12"]),
        ("cupcakes x6", ["cupcake", "x6"]),
        ("alfajores x12", ["alfajor", "x12"]),
        ("trufas x12", ["trufa", "x12"]),
        ("trufas x6", ["trufa"]),
        ("cheesecake corazon", ["cheescake", "corazon"]),
        ("cheesecake corazon", ["cheesecake", "corazon"]),
        ("cheesecake mediano", ["cheesecake", "mediano"]),
        ("cheesecake grande", ["cheesecake", "grande"]),
        ("cheesecake personal", ["cheesecake"]),
        ("brownie mega corazon", ["brownie", "corazon"]),
        ("brownie clasico", ["brownie"]),
        ("cake personal", ["torta", "peque"]),
        ("cake personal", ["torta", "personal"]),
        ("cake 4 a 6", ["torta"]),
        ("desayuno especial", ["desayuno"]),
        ("tres leches x10", ["tres", "leches"]),
        ("tarta vasca", ["vasca"]),
        ("arreglo flores", ["flores"]),
        ("gelatina x4", ["gelatina"]),
        ("sandwich pan", ["sandwich"]),
    ]

    for product_name, tokens in checks:
        if all(token in text_blob for token in tokens):
            match = products_by_name.get(normalize_name(product_name))
            if match:
                return match
    return None


def resolve_sale_total(row: list[Any], product: dict[str, Any] | None) -> float | None:
    candidate_indexes = [10, 9, 8, 7, 5, 11]
    for idx in candidate_indexes:
        if idx < len(row):
            value = num(row[idx])
            if value and value > 0:
                return round(value, 2)
    if product:
        return round(float(product["salePrice"]), 2)
    return None


def extract_sale_rows(path: Path, products_by_name: dict[str, dict[str, Any]]) -> tuple[list[dict[str, Any]], set[str]]:
    book = xlrd.open_workbook(str(path))
    rows: list[dict[str, Any]] = []
    customer_names: set[str] = set()
    for sheet_name in book.sheet_names():
        if sheet_name.upper() == "TOTALES":
            continue
        sheet = book.sheet_by_name(sheet_name)
        for row_idx in range(5, sheet.nrows):
            row = sheet.row_values(row_idx)
            sale_client = text(row[2] if len(row) > 2 else "")
            place = text(row[3] if len(row) > 3 else "")
            observation = text(row[6] if len(row) > 6 else "")
            blob = normalize_spaces(" | ".join(part for part in [sale_client, place, observation] if part))
            if not blob:
                continue
            product = match_product(products_by_name, blob)
            total = resolve_sale_total(row, product)
            if total is None or total <= 0:
                continue

            customer_name = sale_client.title() if probable_customer(sale_client) else None
            if customer_name:
                customer_names.add(customer_name)

            rows.append(
                {
                    "sheet": sheet_name,
                    "date": xls_date(row[1] if len(row) > 1 else None, book.datemode) or sheet_fallback_date(sheet_name, row_idx),
                    "customer_name": customer_name,
                    "description": blob[:200],
                    "notes": normalize_spaces(f"Importado desde ventas {sheet_name}. {blob}")[:500],
                    "product_name": product["name"] if product else "Venta histórica importada",
                    "unit_price": total,
                }
            )
    return rows, customer_names


def import_suppliers(importer: Importer, purchase_rows: list[dict[str, Any]]) -> None:
    for supplier_name in sorted({row["supplier"] for row in purchase_rows}):
        importer.ensure_supplier(supplier_name, "Importado desde compras 2026.")


def import_customers(importer: Importer, customer_names: set[str]) -> None:
    for customer_name in sorted(customer_names):
        importer.ensure_customer(customer_name, "Importado desde ventas 2026.")


def import_purchases(importer: Importer, purchase_rows: list[dict[str, Any]]) -> None:
    ingredient = importer.ingredients_by_name[normalize_name("Compra histórica importada")]
    for row in purchase_rows:
        supplier = importer.suppliers_by_name[normalize_name(row["supplier"])]
        payload = {
            "supplierId": supplier["id"],
            "invoiceNumber": row["invoice"],
            "purchaseDate": row["date"],
            "notes": row["notes"],
            "details": [
                {
                    "ingredientId": ingredient["id"],
                    "description": row["description"],
                    "quantity": 1,
                    "unitOfMeasureId": ingredient["unitOfMeasureId"],
                    "unitCost": row["total"],
                }
            ],
        }
        importer.client.post("purchases", payload)


def import_sales(importer: Importer, sale_rows: list[dict[str, Any]]) -> None:
    for row in sale_rows:
        product = importer.products_by_name[normalize_name(row["product_name"])]
        customer_id = None
        if row["customer_name"]:
            customer_id = importer.customers_by_name[normalize_name(row["customer_name"])]["id"]
        payload = {
            "customerId": customer_id,
            "date": row["date"],
            "notes": row["notes"],
            "paymentMethod": 4,
            "details": [
                {
                    "productId": product["id"],
                    "description": row["description"],
                    "quantity": 1,
                    "unitPrice": row["unit_price"],
                }
            ],
        }
        importer.client.post("sales", payload)


def main() -> int:
    parser = argparse.ArgumentParser(description="Importa datos iniciales de Delicias Bakery a BakeryFlow.")
    parser.add_argument("--api-url", required=True)
    parser.add_argument("--email", required=True)
    parser.add_argument("--password", required=True)
    parser.add_argument("--costs-file", required=True, type=Path)
    parser.add_argument("--purchases-file", required=True, type=Path)
    parser.add_argument("--sales-file", required=True, type=Path)
    args = parser.parse_args()

    client = ApiClient(args.api_url, args.email, args.password)
    importer = Importer(client)
    importer.refresh()
    existing_purchases = client.get("purchases", page=1, pageSize=1).get("totalCount", 0)
    existing_sales = client.get("sales", page=1, pageSize=1).get("totalCount", 0)
    if existing_purchases or existing_sales:
        raise SystemExit(
            f"La base ya tiene datos operativos (compras={existing_purchases}, ventas={existing_sales}). "
            "Detén el proceso y limpia o respalda antes de reimportar."
        )

    seed_units(importer)
    importer.refresh()

    import_products_and_recipes(importer, args.costs_file)
    importer.refresh()

    purchase_rows = extract_purchase_rows(args.purchases_file)
    import_suppliers(importer, purchase_rows)
    importer.refresh()

    sale_rows, customer_names = extract_sale_rows(args.sales_file, importer.products_by_name)
    import_customers(importer, customer_names)
    importer.refresh()

    print(f"Compras a importar: {len(purchase_rows)}")
    import_purchases(importer, purchase_rows)

    print(f"Ventas a importar: {len(sale_rows)}")
    import_sales(importer, sale_rows)

    print("Importación completada.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
